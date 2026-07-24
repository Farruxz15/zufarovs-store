import re
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import Page, sync_playwright


CHANNEL_NAME = "K-shop TOSHKENT OPTOM"

PROFILE_DIR = Path(".telegram_playwright_profile")
OUTPUT_FILE = Path("kshop_catalog.xlsx")
DEBUG_FILE = Path("telegram_debug.html")

MAX_SCROLLS = 1200
STABLE_LIMIT = 35
SCROLL_PAUSE = 1.1


def normalize_text(value: str) -> str:
    return re.sub(r"[ \t]+", " ", value or "").strip()


def extract_price(text: str) -> int | None:
    text = text.replace("\u00a0", " ")
    lines = [normalize_text(line) for line in text.splitlines() if line.strip()]

    explicit_patterns = [
        r"(?i)(?:narxi|цена|price)\s*[:\-]?\s*(\d[\d\s.,]{2,})",
        r"(?i)(\d[\d\s.,]{2,})\s*(?:so['ʻ’`]?m|сўм|сум|uzs)",
    ]

    for line in reversed(lines):
        for pattern in explicit_patterns:
            match = re.search(pattern, line)
            if match:
                digits = re.sub(r"\D", "", match.group(1))
                if digits:
                    value = int(digits)
                    if 1_000 <= value <= 100_000_000:
                        return value

    # Цена часто публикуется отдельной строкой: 120.000 / 120 000
    for line in reversed(lines):
        compact = line.strip()

        if re.fullmatch(r"\d{2,3}(?:[.\s,]\d{3})+", compact):
            value = int(re.sub(r"\D", "", compact))
            if 1_000 <= value <= 100_000_000:
                return value

        if re.fullmatch(r"\d{4,8}", compact):
            value = int(compact)
            if 1_000 <= value <= 100_000_000:
                return value

    return None


def extract_name(text: str) -> str:
    lines = [normalize_text(line) for line in text.splitlines() if line.strip()]

    ignored_patterns = [
        r"^\d{1,2}:\d{2}$",
        r"^\d+\s*(?:views?|просмотр)",
        r"^(?:edited|изменено)$",
        r"^\d{2,3}(?:[.\s,]\d{3})+$",
        r"^\d{4,8}$",
    ]

    for line in lines:
        if len(line) < 3:
            continue

        if any(re.search(pattern, line, re.I) for pattern in ignored_patterns):
            continue

        return line[:300]

    return "Без названия"


def classify_product(text: str) -> str:
    lower = text.lower()

    categories = [
        ("SPF и солнцезащита", ["spf", "sun cream", "sunscreen", "quyosh"]),
        ("Крем для лица", ["cream", "крем", "krem", "moistur"]),
        ("Сыворотка", ["serum", "сыворот", "zardob"]),
        ("Тонер", ["toner", "тонер"]),
        ("Очищение", ["cleanser", "cleansing", "пенка", "yuvish", "gel"]),
        ("Маска", ["mask", "маска", "niqob"]),
        ("Шампунь", ["shampoo", "шампун", "shampun"]),
        ("Уход за волосами", ["hair", "волос", "soch"]),
        ("Уход за телом", ["body", "тело", "tana"]),
        ("Декоративная косметика", ["lip", "помада", "cushion", "foundation"]),
    ]

    for category, keywords in categories:
        if any(keyword in lower for keyword in keywords):
            return category

    return "Другое"


def detect_scroll_container(page: Page) -> Any:
    result = page.evaluate(
        """
        () => {
            const elements = [...document.querySelectorAll('*')];

            const candidates = elements
                .map((el, index) => {
                    const style = getComputedStyle(el);
                    const rect = el.getBoundingClientRect();

                    return {
                        index,
                        el,
                        scrollHeight: el.scrollHeight,
                        clientHeight: el.clientHeight,
                        overflowY: style.overflowY,
                        width: rect.width,
                        height: rect.height,
                        area: rect.width * rect.height
                    };
                })
                .filter(item =>
                    item.clientHeight > 300 &&
                    item.scrollHeight > item.clientHeight + 200 &&
                    item.width > 400 &&
                    ['auto', 'scroll', 'overlay'].includes(item.overflowY)
                )
                .sort((a, b) => b.area - a.area);

            if (!candidates.length) return null;

            const target = candidates[0].el;
            target.setAttribute('data-playwright-scroll-container', 'true');

            return {
                tag: target.tagName,
                className: String(target.className || ''),
                scrollHeight: target.scrollHeight,
                clientHeight: target.clientHeight,
                scrollTop: target.scrollTop
            };
        }
        """
    )

    if not result:
        return None

    return page.locator('[data-playwright-scroll-container="true"]')


def collect_visible_messages(page: Page) -> list[dict[str, Any]]:
    return page.evaluate(
        """
        () => {
            const selectors = [
                '[data-message-id]',
                '[data-mid]',
                '.Message',
                '.message',
                '.message-list-item',
                '.bubble',
                '.bubble-content',
                'article'
            ];

            const all = [];
            const seenNodes = new Set();

            for (const selector of selectors) {
                for (const element of document.querySelectorAll(selector)) {
                    if (seenNodes.has(element)) continue;
                    seenNodes.add(element);

                    const rect = element.getBoundingClientRect();
                    const text = (element.innerText || '').trim();

                    if (!text || text.length < 3) continue;
                    if (rect.width < 250 || rect.height < 20) continue;
                    if (rect.bottom < 0 || rect.top > window.innerHeight) continue;

                    const id =
                        element.getAttribute('data-message-id') ||
                        element.getAttribute('data-mid') ||
                        element.id ||
                        '';

                    const timeElement = element.querySelector('time');
                    const date =
                        timeElement?.getAttribute('datetime') ||
                        timeElement?.getAttribute('title') ||
                        timeElement?.innerText ||
                        '';

                    const imageUrls = [...element.querySelectorAll('img')]
                        .map(img => img.currentSrc || img.src || '')
                        .filter(src =>
                            src &&
                            !src.startsWith('data:image/svg') &&
                            !src.includes('emoji')
                        );

                    all.push({
                        id,
                        text,
                        date,
                        image_urls: [...new Set(imageUrls)],
                        top: rect.top,
                        class_name: String(element.className || '')
                    });
                }
            }

            // Удаляем вложенные элементы с одинаковым или почти одинаковым текстом
            all.sort((a, b) => a.text.length - b.text.length);

            const cleaned = [];

            for (const item of all) {
                const duplicate = cleaned.some(existing =>
                    existing.text === item.text ||
                    (
                        existing.text.length > 30 &&
                        item.text.includes(existing.text) &&
                        item.text.length < existing.text.length * 1.15
                    )
                );

                if (!duplicate) cleaned.push(item);
            }

            return cleaned;
        }
        """
    )


def save_excel(records: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "K-shop ассортимент"

    headers = [
        "№",
        "Категория",
        "Название товара",
        "Цена, сум",
        "Дата",
        "Полный текст объявления",
        "ID сообщения",
        "Ссылки на изображения",
    ]

    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sorted_records = sorted(
        records,
        key=lambda item: (
            item.get("category", ""),
            item.get("name", ""),
            item.get("text", ""),
        ),
    )

    for number, item in enumerate(sorted_records, start=1):
        sheet.append(
            [
                number,
                item["category"],
                item["name"],
                item["price"],
                item.get("date", ""),
                item["text"],
                item.get("id", ""),
                "\n".join(item.get("image_urls", [])),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 7,
        "B": 25,
        "C": 42,
        "D": 18,
        "E": 20,
        "F": 80,
        "G": 20,
        "H": 55,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if row[3].value:
            row[3].number_format = '#,##0" сум"'

    workbook.save(OUTPUT_FILE)


def main() -> None:
    PROFILE_DIR.mkdir(exist_ok=True)

    records: dict[str, dict[str, Any]] = {}

    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR.resolve()),
            headless=False,
            viewport=None,
            args=["--start-maximized"],
        )

        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://web.telegram.org/a/", wait_until="domcontentloaded")
        page.wait_for_timeout(3000)

        print("\nОткрой канал K-shop TOSHKENT OPTOM в окне Chromium.")
        print("Убедись, что справа видны сообщения с товарами.")
        input("После этого вернись в Terminal и нажми Enter... ")

        page.wait_for_timeout(2000)

        scroll_container = detect_scroll_container(page)

        if scroll_container is None:
            DEBUG_FILE.write_text(page.content(), encoding="utf-8")
            print("\nНе найден контейнер сообщений.")
            print(f"Диагностика сохранена: {DEBUG_FILE.resolve()}")
            input("Нажми Enter для закрытия браузера... ")
            context.close()
            return

        print("\nКонтейнер сообщений найден.")
        print("Начинаю сбор и прокрутку к самым старым публикациям.\n")

        stable_rounds = 0
        previous_total = 0

        for step in range(1, MAX_SCROLLS + 1):
            visible = collect_visible_messages(page)

            for message in visible:
                text = normalize_text(message.get("text", ""))

                if not text:
                    continue

                # Более устойчивый ключ, когда Telegram не показывает ID
                key = message.get("id") or re.sub(
                    r"\W+",
                    "",
                    text.lower(),
                )[:250]

                if not key:
                    continue

                price = extract_price(text)

                records[key] = {
                    "id": message.get("id", ""),
                    "date": message.get("date", ""),
                    "text": text,
                    "name": extract_name(text),
                    "price": price,
                    "category": classify_product(text),
                    "image_urls": message.get("image_urls", []),
                }

            current_total = len(records)

            if current_total == previous_total:
                stable_rounds += 1
            else:
                stable_rounds = 0
                previous_total = current_total

            if step == 1 or step % 10 == 0:
                print(
                    f"Шаг {step:4d} | "
                    f"видимых блоков: {len(visible):3d} | "
                    f"уникальных записей: {current_total:4d}"
                )

            position = scroll_container.evaluate(
                """
                element => ({
                    top: element.scrollTop,
                    height: element.scrollHeight,
                    client: element.clientHeight
                })
                """
            )

            # Telegram может иметь нормальное или обратное направление прокрутки
            scroll_container.evaluate(
                """
                element => {
                    const amount = Math.max(700, element.clientHeight * 0.82);
                    element.scrollBy(0, -amount);
                    element.dispatchEvent(new Event('scroll', { bubbles: true }));
                }
                """
            )

            page.wait_for_timeout(int(SCROLL_PAUSE * 1000))

            new_position = scroll_container.evaluate(
                """
                element => ({
                    top: element.scrollTop,
                    height: element.scrollHeight,
                    client: element.clientHeight
                })
                """
            )

            if (
                position["top"] == new_position["top"]
                and stable_rounds >= 8
            ):
                # Дополнительная попытка через колесо мыши
                box = scroll_container.bounding_box()

                if box:
                    page.mouse.move(
                        box["x"] + box["width"] / 2,
                        box["y"] + box["height"] / 2,
                    )
                    page.mouse.wheel(0, -4000)
                    page.wait_for_timeout(1500)

            if stable_rounds >= STABLE_LIMIT:
                print(
                    "\nНовые сообщения долго не появляются. "
                    "Считаю прокрутку завершённой."
                )
                break

        # Последний сбор видимой области
        for message in collect_visible_messages(page):
            text = normalize_text(message.get("text", ""))

            if not text:
                continue

            key = message.get("id") or re.sub(
                r"\W+",
                "",
                text.lower(),
            )[:250]

            records[key] = {
                "id": message.get("id", ""),
                "date": message.get("date", ""),
                "text": text,
                "name": extract_name(text),
                "price": extract_price(text),
                "category": classify_product(text),
                "image_urls": message.get("image_urls", []),
            }

        final_records = list(records.values())

        save_excel(final_records)
        DEBUG_FILE.write_text(page.content(), encoding="utf-8")

        priced_count = sum(
            1 for item in final_records if item.get("price") is not None
        )

        print("\nГотово.")
        print(f"Всего уникальных записей: {len(final_records)}")
        print(f"Записей с распознанной ценой: {priced_count}")
        print(f"Excel: {OUTPUT_FILE.resolve()}")
        print(f"Диагностический HTML: {DEBUG_FILE.resolve()}")

        input("\nНажми Enter, чтобы закрыть браузер... ")
        context.close()


if __name__ == "__main__":
    main()
PYcat > telegram_catalog.py <<'PY'
import re
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import Page, sync_playwright


CHANNEL_NAME = "K-shop TOSHKENT OPTOM"

PROFILE_DIR = Path(".telegram_playwright_profile")
OUTPUT_FILE = Path("kshop_catalog.xlsx")
DEBUG_FILE = Path("telegram_debug.html")

MAX_SCROLLS = 1200
STABLE_LIMIT = 35
SCROLL_PAUSE = 1.1


def normalize_text(value: str) -> str:
    return re.sub(r"[ \t]+", " ", value or "").strip()


def extract_price(text: str) -> int | None:
    text = text.replace("\u00a0", " ")
    lines = [normalize_text(line) for line in text.splitlines() if line.strip()]

    explicit_patterns = [
        r"(?i)(?:narxi|цена|price)\s*[:\-]?\s*(\d[\d\s.,]{2,})",
        r"(?i)(\d[\d\s.,]{2,})\s*(?:so['ʻ’`]?m|сўм|сум|uzs)",
    ]

    for line in reversed(lines):
        for pattern in explicit_patterns:
            match = re.search(pattern, line)
            if match:
                digits = re.sub(r"\D", "", match.group(1))
                if digits:
                    value = int(digits)
                    if 1_000 <= value <= 100_000_000:
                        return value

    # Цена часто публикуется отдельной строкой: 120.000 / 120 000
    for line in reversed(lines):
        compact = line.strip()

        if re.fullmatch(r"\d{2,3}(?:[.\s,]\d{3})+", compact):
            value = int(re.sub(r"\D", "", compact))
            if 1_000 <= value <= 100_000_000:
                return value

        if re.fullmatch(r"\d{4,8}", compact):
            value = int(compact)
            if 1_000 <= value <= 100_000_000:
                return value

    return None


def extract_name(text: str) -> str:
    lines = [normalize_text(line) for line in text.splitlines() if line.strip()]

    ignored_patterns = [
        r"^\d{1,2}:\d{2}$",
        r"^\d+\s*(?:views?|просмотр)",
        r"^(?:edited|изменено)$",
        r"^\d{2,3}(?:[.\s,]\d{3})+$",
        r"^\d{4,8}$",
    ]

    for line in lines:
        if len(line) < 3:
            continue

        if any(re.search(pattern, line, re.I) for pattern in ignored_patterns):
            continue

        return line[:300]

    return "Без названия"


def classify_product(text: str) -> str:
    lower = text.lower()

    categories = [
        ("SPF и солнцезащита", ["spf", "sun cream", "sunscreen", "quyosh"]),
        ("Крем для лица", ["cream", "крем", "krem", "moistur"]),
        ("Сыворотка", ["serum", "сыворот", "zardob"]),
        ("Тонер", ["toner", "тонер"]),
        ("Очищение", ["cleanser", "cleansing", "пенка", "yuvish", "gel"]),
        ("Маска", ["mask", "маска", "niqob"]),
        ("Шампунь", ["shampoo", "шампун", "shampun"]),
        ("Уход за волосами", ["hair", "волос", "soch"]),
        ("Уход за телом", ["body", "тело", "tana"]),
        ("Декоративная косметика", ["lip", "помада", "cushion", "foundation"]),
    ]

    for category, keywords in categories:
        if any(keyword in lower for keyword in keywords):
            return category

    return "Другое"


def detect_scroll_container(page: Page) -> Any:
    result = page.evaluate(
        """
        () => {
            const elements = [...document.querySelectorAll('*')];

            const candidates = elements
                .map((el, index) => {
                    const style = getComputedStyle(el);
                    const rect = el.getBoundingClientRect();

                    return {
                        index,
                        el,
                        scrollHeight: el.scrollHeight,
                        clientHeight: el.clientHeight,
                        overflowY: style.overflowY,
                        width: rect.width,
                        height: rect.height,
                        area: rect.width * rect.height
                    };
                })
                .filter(item =>
                    item.clientHeight > 300 &&
                    item.scrollHeight > item.clientHeight + 200 &&
                    item.width > 400 &&
                    ['auto', 'scroll', 'overlay'].includes(item.overflowY)
                )
                .sort((a, b) => b.area - a.area);

            if (!candidates.length) return null;

            const target = candidates[0].el;
            target.setAttribute('data-playwright-scroll-container', 'true');

            return {
                tag: target.tagName,
                className: String(target.className || ''),
                scrollHeight: target.scrollHeight,
                clientHeight: target.clientHeight,
                scrollTop: target.scrollTop
            };
        }
        """
    )

    if not result:
        return None

    return page.locator('[data-playwright-scroll-container="true"]')


def collect_visible_messages(page: Page) -> list[dict[str, Any]]:
    return page.evaluate(
        """
        () => {
            const selectors = [
                '[data-message-id]',
                '[data-mid]',
                '.Message',
                '.message',
                '.message-list-item',
                '.bubble',
                '.bubble-content',
                'article'
            ];

            const all = [];
            const seenNodes = new Set();

            for (const selector of selectors) {
                for (const element of document.querySelectorAll(selector)) {
                    if (seenNodes.has(element)) continue;
                    seenNodes.add(element);

                    const rect = element.getBoundingClientRect();
                    const text = (element.innerText || '').trim();

                    if (!text || text.length < 3) continue;
                    if (rect.width < 250 || rect.height < 20) continue;
                    if (rect.bottom < 0 || rect.top > window.innerHeight) continue;

                    const id =
                        element.getAttribute('data-message-id') ||
                        element.getAttribute('data-mid') ||
                        element.id ||
                        '';

                    const timeElement = element.querySelector('time');
                    const date =
                        timeElement?.getAttribute('datetime') ||
                        timeElement?.getAttribute('title') ||
                        timeElement?.innerText ||
                        '';

                    const imageUrls = [...element.querySelectorAll('img')]
                        .map(img => img.currentSrc || img.src || '')
                        .filter(src =>
                            src &&
                            !src.startsWith('data:image/svg') &&
                            !src.includes('emoji')
                        );

                    all.push({
                        id,
                        text,
                        date,
                        image_urls: [...new Set(imageUrls)],
                        top: rect.top,
                        class_name: String(element.className || '')
                    });
                }
            }

            // Удаляем вложенные элементы с одинаковым или почти одинаковым текстом
            all.sort((a, b) => a.text.length - b.text.length);

            const cleaned = [];

            for (const item of all) {
                const duplicate = cleaned.some(existing =>
                    existing.text === item.text ||
                    (
                        existing.text.length > 30 &&
                        item.text.includes(existing.text) &&
                        item.text.length < existing.text.length * 1.15
                    )
                );

                if (!duplicate) cleaned.push(item);
            }

            return cleaned;
        }
        """
    )


def save_excel(records: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "K-shop ассортимент"

    headers = [
        "№",
        "Категория",
        "Название товара",
        "Цена, сум",
        "Дата",
        "Полный текст объявления",
        "ID сообщения",
        "Ссылки на изображения",
    ]

    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sorted_records = sorted(
        records,
        key=lambda item: (
            item.get("category", ""),
            item.get("name", ""),
            item.get("text", ""),
        ),
    )

    for number, item in enumerate(sorted_records, start=1):
        sheet.append(
            [
                number,
                item["category"],
                item["name"],
                item["price"],
                item.get("date", ""),
                item["text"],
                item.get("id", ""),
                "\n".join(item.get("image_urls", [])),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 7,
        "B": 25,
        "C": 42,
        "D": 18,
        "E": 20,
        "F": 80,
        "G": 20,
        "H": 55,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if row[3].value:
            row[3].number_format = '#,##0" сум"'

    workbook.save(OUTPUT_FILE)


def main() -> None:
    PROFILE_DIR.mkdir(exist_ok=True)

    records: dict[str, dict[str, Any]] = {}

    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR.resolve()),
            headless=False,
            viewport=None,
            args=["--start-maximized"],
        )

        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://web.telegram.org/a/", wait_until="domcontentloaded")
        page.wait_for_timeout(3000)

        print("\nОткрой канал K-shop TOSHKENT OPTOM в окне Chromium.")
        print("Убедись, что справа видны сообщения с товарами.")
        input("После этого вернись в Terminal и нажми Enter... ")

        page.wait_for_timeout(2000)

        scroll_container = detect_scroll_container(page)

        if scroll_container is None:
            DEBUG_FILE.write_text(page.content(), encoding="utf-8")
            print("\nНе найден контейнер сообщений.")
            print(f"Диагностика сохранена: {DEBUG_FILE.resolve()}")
            input("Нажми Enter для закрытия браузера... ")
            context.close()
            return

        print("\nКонтейнер сообщений найден.")
        print("Начинаю сбор и прокрутку к самым старым публикациям.\n")

        stable_rounds = 0
        previous_total = 0

        for step in range(1, MAX_SCROLLS + 1):
            visible = collect_visible_messages(page)

            for message in visible:
                text = normalize_text(message.get("text", ""))

                if not text:
                    continue

                # Более устойчивый ключ, когда Telegram не показывает ID
                key = message.get("id") or re.sub(
                    r"\W+",
                    "",
                    text.lower(),
                )[:250]

                if not key:
                    continue

                price = extract_price(text)

                records[key] = {
                    "id": message.get("id", ""),
                    "date": message.get("date", ""),
                    "text": text,
                    "name": extract_name(text),
                    "price": price,
                    "category": classify_product(text),
                    "image_urls": message.get("image_urls", []),
                }

            current_total = len(records)

            if current_total == previous_total:
                stable_rounds += 1
            else:
                stable_rounds = 0
                previous_total = current_total

            if step == 1 or step % 10 == 0:
                print(
                    f"Шаг {step:4d} | "
                    f"видимых блоков: {len(visible):3d} | "
                    f"уникальных записей: {current_total:4d}"
                )

            position = scroll_container.evaluate(
                """
                element => ({
                    top: element.scrollTop,
                    height: element.scrollHeight,
                    client: element.clientHeight
                })
                """
            )

            # Telegram может иметь нормальное или обратное направление прокрутки
            scroll_container.evaluate(
                """
                element => {
                    const amount = Math.max(700, element.clientHeight * 0.82);
                    element.scrollBy(0, -amount);
                    element.dispatchEvent(new Event('scroll', { bubbles: true }));
                }
                """
            )

            page.wait_for_timeout(int(SCROLL_PAUSE * 1000))

            new_position = scroll_container.evaluate(
                """
                element => ({
                    top: element.scrollTop,
                    height: element.scrollHeight,
                    client: element.clientHeight
                })
                """
            )

            if (
                position["top"] == new_position["top"]
                and stable_rounds >= 8
            ):
                # Дополнительная попытка через колесо мыши
                box = scroll_container.bounding_box()

                if box:
                    page.mouse.move(
                        box["x"] + box["width"] / 2,
                        box["y"] + box["height"] / 2,
                    )
                    page.mouse.wheel(0, -4000)
                    page.wait_for_timeout(1500)

            if stable_rounds >= STABLE_LIMIT:
                print(
                    "\nНовые сообщения долго не появляются. "
                    "Считаю прокрутку завершённой."
                )
                break

        # Последний сбор видимой области
        for message in collect_visible_messages(page):
            text = normalize_text(message.get("text", ""))

            if not text:
                continue

            key = message.get("id") or re.sub(
                r"\W+",
                "",
                text.lower(),
            )[:250]

            records[key] = {
                "id": message.get("id", ""),
                "date": message.get("date", ""),
                "text": text,
                "name": extract_name(text),
                "price": extract_price(text),
                "category": classify_product(text),
                "image_urls": message.get("image_urls", []),
            }

        final_records = list(records.values())

        save_excel(final_records)
        DEBUG_FILE.write_text(page.content(), encoding="utf-8")

        priced_count = sum(
            1 for item in final_records if item.get("price") is not None
        )

        print("\nГотово.")
        print(f"Всего уникальных записей: {len(final_records)}")
        print(f"Записей с распознанной ценой: {priced_count}")
        print(f"Excel: {OUTPUT_FILE.resolve()}")
        print(f"Диагностический HTML: {DEBUG_FILE.resolve()}")

        input("\nНажми Enter, чтобы закрыть браузер... ")
        context.close()


if __name__ == "__main__":
    main()
