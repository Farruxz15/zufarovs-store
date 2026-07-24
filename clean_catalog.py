import re
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SOURCE = Path("kshop_catalog.xlsx")
OUTPUT = Path("kshop_catalog_clean.xlsx")


def normalize(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_text(value):
    text = normalize(value).lower()

    # Удаляем время, просмотры и служебные фрагменты Telegram
    text = re.sub(r"\b\d{1,2}:\d{2}\b", " ", text)
    text = re.sub(r"\b\d+\s*(views?|просмотров?|просмотра)\b", " ", text)
    text = re.sub(r"\b(edited|изменено)\b", " ", text)

    # Нормализуем разные разделители в ценах и тексте
    text = text.replace("’", "'").replace("ʻ", "'")
    text = re.sub(r"[^\wа-яёўқғҳ\s]", " ", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip()

    return text


def clean_name(value):
    name = clean_text(value)

    # Удаляем возможную цену из названия
    name = re.sub(r"\b\d{4,8}\b", " ", name)
    name = re.sub(r"\s+", " ", name).strip()

    return name


def is_valid_product(name, text, price):
    if not isinstance(price, (int, float)):
        return False

    if not 1_000 <= price <= 100_000_000:
        return False

    combined = f"{name} {text}".lower()

    excluded = [
        "unread messages",
        "архив чатов",
        "search",
        "поиск",
        "subscribers",
        "подписчик",
        "pinned message",
        "закрепленное сообщение",
    ]

    if any(item in combined for item in excluded):
        return False

    return len(normalize(text)) >= 5


def product_key(name, text, price):
    normalized_name = clean_name(name)
    normalized_text = clean_text(text)

    # Берём первые информативные части текста.
    # Это убирает дубли, где отличаются только время, просмотры или оформление.
    text_words = normalized_text.split()
    compact_text = " ".join(text_words[:45])

    return (
        normalized_name[:150],
        compact_text,
        int(price),
    )


source_wb = load_workbook(SOURCE)
source_ws = source_wb.active

output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "Без дубликатов"

headers = [
    "№",
    "Категория",
    "Название товара",
    "Цена, сум",
    "Дата",
    "Полный текст объявления",
    "ID сообщения",
    "Ссылки на изображения",
]

output_ws.append(headers)

header_fill = PatternFill("solid", fgColor="D9EAF7")

for cell in output_ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

unique_products = {}
duplicate_count = 0

for row in source_ws.iter_rows(min_row=2, values_only=True):
    _, category, name, price, date, text, message_id, image_urls = row

    name = normalize(name)
    text = normalize(text)
    category = normalize(category) or "Другое"
    date = normalize(date)
    message_id = normalize(message_id)
    image_urls = normalize(image_urls)

    if not is_valid_product(name, text, price):
        continue

    key = product_key(name, text, price)

    if key in unique_products:
        duplicate_count += 1

        # Сохраняем более полную запись
        old = unique_products[key]
        if len(text) > len(old["text"]):
            unique_products[key] = {
                "category": category,
                "name": name,
                "price": int(price),
                "date": date,
                "text": text,
                "message_id": message_id,
                "image_urls": image_urls,
            }
        continue

    unique_products[key] = {
        "category": category,
        "name": name,
        "price": int(price),
        "date": date,
        "text": text,
        "message_id": message_id,
        "image_urls": image_urls,
    }

clean_rows = list(unique_products.values())

clean_rows.sort(
    key=lambda item: (
        item["category"].lower(),
        item["name"].lower(),
        item["price"],
    )
)

for number, item in enumerate(clean_rows, start=1):
    output_ws.append(
        [
            number,
            item["category"],
            item["name"],
            item["price"],
            item["date"],
            item["text"],
            item["message_id"],
            item["image_urls"],
        ]
    )

output_ws.freeze_panes = "A2"
output_ws.auto_filter.ref = output_ws.dimensions

widths = {
    "A": 7,
    "B": 25,
    "C": 42,
    "D": 18,
    "E": 20,
    "F": 80,
    "G": 20,
    "H": 55,
}

for column, width in widths.items():
    output_ws.column_dimensions[column].width = width

for row in output_ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    row[3].number_format = '#,##0" сум"'

output_wb.save(OUTPUT)

print(f"Готово: {OUTPUT.resolve()}")
print(f"Уникальных товаров: {len(clean_rows)}")
print(f"Удалено дубликатов: {duplicate_count}")
